"""
Synthesize statistics report for a generation.
"""

from pathlib import Path
from functools import cached_property
from dataclasses import dataclass
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from parse import parse
from loguru import logger
import re

from src.preprocessor.api import APICollection
from src.generator.worker import GenerationState
from src import vars as global_vars


class FunctionReport:
    """
    API function report.
    """

    @dataclass
    class FunctionEntry:
        """
        An API function data entry.
        """

        name: str
        """
        Function name.
        """

        location: str
        """
        Function location.
        """

        header: str
        """
        The header file where the function is declared.
        """

        is_tested: bool
        """
        Whether the function is tested.
        """

        is_deprecated: bool
        """
        Whether the function is deprecated.
        """

        occurences: float
        """
        Number of times the function is used.
        """

        failed_times: float
        """
        Number of times the function failed to generate.
        """

    @cached_property
    def total_count(self) -> int:
        """
        The total count of functions.
        """
        return len(self.entries)

    @cached_property
    def tested_count(self) -> int:
        """
        The count of tested functions.
        """
        return sum(row.is_tested for row in self.entries)

    @cached_property
    def untested_count(self) -> int:
        """
        The count of untested functions.
        """
        return self.total_count - self.tested_count

    @cached_property
    def function_coverage(self) -> float:
        """
        The function coverage percentage.
        """
        return (self.tested_count / self.total_count) if self.total_count else 0

    @cached_property
    def total_occurences(self) -> int:
        """
        The total number of function occurences.
        """
        return sum(row.occurences for row in self.entries)

    @cached_property
    def average_occurences(self) -> float:
        """
        The average number of function occurences.
        """
        return (self.total_occurences / self.total_count) if self.total_count else 0

    @cached_property
    def average_failed_times(self) -> float:
        """
        The average number of failed times.
        """
        return (
            (sum(row.failed_times for row in self.entries) / self.total_count)
            if self.total_count
            else 0
        )

    def add_to_workbook(self, workbook: Workbook):
        """
        Add the function report to a workbook as a new sheet.

        :param workbook: The workbook.
        """
        # Create a new sheet
        function_sheet = workbook.create_sheet("API Functions")

        # Write the header
        (
            function_sheet["A1"],
            function_sheet["B1"],
            function_sheet["C1"],
            function_sheet["D1"],
            function_sheet["E1"],
            function_sheet["F1"],
            function_sheet["G1"],
        ) = (
            "Function Name",
            "Location",
            "Header File",
            "isTested",
            "isDeprecated",
            "Occurences in Fuzz Driver",
            "Failed Times",
        )

        # Write the data
        for i, row in enumerate(self.entries, start=2):
            (
                function_sheet[f"A{i}"],
                function_sheet[f"B{i}"],
                function_sheet[f"C{i}"],
                function_sheet[f"D{i}"],
                function_sheet[f"E{i}"],
                function_sheet[f"F{i}"],
                function_sheet[f"G{i}"],
            ) = (
                row.name,
                row.location,
                row.header,
                row.is_tested,
                row.is_deprecated,
                row.occurences,
                row.failed_times,
            )

        # Write the statistics
        cur_row = i + 1
        function_sheet[f"A{cur_row}"] = "————————————————"

        def _write_stat(name, value):
            nonlocal cur_row
            cur_row += 1
            function_sheet[f"A{cur_row}"], function_sheet[f"B{cur_row}"] = name, value

        _write_stat("Total Count", self.total_count)
        _write_stat("Tested Count", self.tested_count)
        _write_stat("Untested Count", self.untested_count)
        _write_stat("Function Coverage", f"{self.function_coverage * 100:.2f}%")
        _write_stat("Total Occurences", self.total_occurences)
        _write_stat("Average Occurences", self.average_occurences)
        _write_stat("Average Failed Times", self.average_failed_times)

        # set styles
        # bold the header row
        bold = Font(bold=True, name="Arial")
        for cell in function_sheet["1:1"]:
            cell.font = bold
        # bold the A column
        for cell in function_sheet["A:A"]:
            cell.font = bold
        # set column width
        width = {
            "A": 35,
            "B": 100,
            "C": 100,
            "D": 9,
            "E": 13,
            "F": 11,
            "G": 13,
        }
        for col, w in width.items():
            function_sheet.column_dimensions[col].width = w

    def __init__(self, function_entries: list[FunctionEntry]):
        """
        Initialize the function report.

        :param function_entries: The function entries.
        """
        self.entries = function_entries

    @classmethod
    def from_collected_data(
        cls, api_collection: APICollection, gen_state: GenerationState
    ):
        """
        Create a function report from collected data.

        :param api_collection: The API collection.
        :param gen_state: The generation state.
        """
        # function entries dict, key is function location, value is FunctionEntry
        # use dict but not list to quickly look up and update
        function_entries: dict[str, FunctionReport.FunctionEntry] = dict()

        # get basic data from api collection
        for func in api_collection.funcs:
            new_entry = cls.FunctionEntry(
                name=func.name,
                location=func.loc,
                header=func.header,
                is_tested=False,
                is_deprecated=False,
                occurences=0,
                failed_times=0,
            )
            function_entries[func.loc] = new_entry

        # get additional data from generation state
        for (
            func,
            occurences,
        ) in gen_state.scheduler_statistics.function_occurrences.items():
            function_entries[func.loc].occurences = occurences

        for func, is_tested in gen_state.scheduler_statistics.function_tested.items():
            function_entries[func.loc].is_tested = is_tested

        for (
            func,
            failed_times,
        ) in gen_state.scheduler_statistics.function_failed_times.items():
            function_entries[func.loc].failed_times = failed_times

        for func in gen_state.scheduler_statistics.deprecated_functions:
            function_entries[func.loc].is_deprecated = True

        return cls(list(function_entries.values()))


class DriverReport:
    """
    Fuzz driver report.
    """

    @dataclass
    class DriverEntry:
        """
        A fuzz driver data entry.
        """

        id: int
        """
        The driver ID.
        """

        is_success: bool
        """
        Whether the driver is successfully generated and saved.
        """

        query_count: int
        """
        The number of queries in the driver.
        """

        target_functions: list[str]
        """
        The target functions in the driver.
        """

        @cached_property
        def target_function_count(self) -> int:
            """
            The count of target functions.
            """
            return len(self.target_functions)

    @cached_property
    def total_count(self) -> int:
        """
        The total count of drivers.
        """
        return len(self.entries)

    @cached_property
    def success_count(self) -> int:
        """
        The count of successful drivers.
        """
        return sum(row.is_success for row in self.entries)

    @cached_property
    def failure_count(self) -> int:
        """
        The count of failed drivers.
        """
        return self.total_count - self.success_count

    @cached_property
    def success_rate(self) -> float:
        """
        The success rate of drivers.
        """
        return (self.success_count / self.total_count) if self.total_count else 0

    @cached_property
    def total_queries(self) -> int:
        """
        The total number of queries in drivers.
        """
        return sum(row.query_count for row in self.entries)

    @cached_property
    def average_queries(self) -> float:
        """
        The average number of queries in drivers.
        """
        return (self.total_queries / self.total_count) if self.total_count else 0

    @cached_property
    def average_queries_per_usable_driver(self) -> float:
        """
        The average number of queries per usable driver.
        """
        return (self.total_queries / self.success_count) if self.success_count else 0

    @cached_property
    def average_tokens(self) -> tuple[float, float]:
        """
        The average number of tokens consumed.
        """
        return (
            tuple(x / self.total_count for x in self.total_tokens)
            if self.total_count
            else (0, 0)
        )

    @cached_property
    def average_tokens_per_usable_driver(self) -> tuple[float, float]:
        """
        The average number of tokens consumed per usable driver.
        """
        return (
            tuple(x / self.success_count for x in self.total_tokens)
            if self.success_count
            else (0, 0)
        )

    @cached_property
    def average_time(self) -> float:
        """
        The average time consumed.
        """
        return (self.total_time / self.total_count) if self.total_count else 0

    @cached_property
    def average_time_per_usable_driver(self) -> float:
        """
        The average time consumed per usable driver.
        """
        return (self.total_time / self.success_count) if self.success_count else 0

    @cached_property
    def total_target_functions(self) -> int:
        """
        The total number of target functions.
        """
        return sum(row.target_function_count for row in self.entries)

    @cached_property
    def average_target_functions(self) -> float:
        """
        The average number of target functions.
        """
        return (
            (self.total_target_functions / self.success_count)
            if self.success_count
            else 0
        )

    def add_to_workbook(self, workbook: Workbook):
        """
        Add the driver report to a workbook as a new sheet.

        :param workbook: The workbook.
        """
        # Create a new sheet
        driver_sheet = workbook.create_sheet("Fuzz Drivers")

        # Write the header
        (
            driver_sheet["A1"],
            driver_sheet["B1"],
            driver_sheet["C1"],
            driver_sheet["D1"],
            driver_sheet["E1"],
        ) = (
            "Driver ID",
            "isSuccess",
            "Query Count",
            "Target Function Count",
            "Target Functions",
        )
        # bold the header row
        bold = Font(bold=True)
        for cell in driver_sheet["1:1"]:
            cell.font = bold

        # Write the data
        for i, row in enumerate(self.entries, start=2):
            (
                driver_sheet[f"A{i}"],
                driver_sheet[f"B{i}"],
                driver_sheet[f"C{i}"],
                driver_sheet[f"D{i}"],
                driver_sheet[f"E{i}"],
            ) = (
                row.id,
                row.is_success,
                row.query_count,
                row.target_function_count,
                "; ".join(row.target_functions),
            )

        # Write the statistics
        cur_row = i + 1
        driver_sheet[f"A{cur_row}"] = "————————————————"

        def _write_stat(name, value):
            nonlocal cur_row
            cur_row += 1
            driver_sheet[f"A{cur_row}"], driver_sheet[f"B{cur_row}"] = name, value

        _write_stat("Total Count", self.total_count)
        _write_stat("Success Count", self.success_count)
        _write_stat("Failure Count", self.failure_count)
        _write_stat("Success Rate", f"{self.success_rate * 100:.2f}%")
        _write_stat("Total Queries", self.total_queries)
        _write_stat("Average Queries", self.average_queries)
        _write_stat(
            "Average Queries per Usable Driver", self.average_queries_per_usable_driver
        )
        _write_stat("Total Tokens", ", ".join(map(str, self.total_tokens)))
        _write_stat(
            "Average Tokens", ", ".join(map(lambda x: f"{x:.2f}", self.average_tokens))
        )
        _write_stat(
            "Average Tokens per Usable Driver",
            ", ".join(map(lambda x: f"{x:.2f}", self.average_tokens_per_usable_driver)),
        )
        _write_stat("Total Time", self.total_time)
        _write_stat("Average Time", self.average_time)
        _write_stat(
            "Average Time per Usable Driver", self.average_time_per_usable_driver
        )
        _write_stat("Total Target Functions", self.total_target_functions)
        _write_stat("Average Target Functions", self.average_target_functions)

        # set styles
        # bold the header row
        bold = Font(bold=True, name="Arial")
        for cell in driver_sheet["1:1"]:
            cell.font = bold
        # bold the A column
        for cell in driver_sheet["A:A"]:
            cell.font = bold
        # set column width
        width = {
            "A": 31,
            "B": 28,
            "C": 12,
            "D": 22,
            "E": 200,
        }
        for col, w in width.items():
            driver_sheet.column_dimensions[col].width = w

    def __init__(
        self,
        driver_entries: list[DriverEntry],
        total_tokens: tuple[int, int],
        total_time: float,
    ):
        """
        Initialize the driver report.

        :param driver_entries: The driver entries.
        :param total_tokens: The total number of tokens consumed.
        :param total_time: The total time consumed.
        """
        self.entries = driver_entries
        self.total_tokens = total_tokens
        self.total_time = total_time

    @classmethod
    def from_collected_data(
        cls,
        gen_state: GenerationState,
        driver_out_path: Path,
    ):
        """
        Create a driver report from collected data.

        :param gen_state: The generation state.
        :param driver_out_path: The driver output path.
        """
        # driver entries dict, key is driver id, value is DriverEntry
        # use dict but not list to quickly look up and update
        driver_entries: dict[int, DriverReport.DriverEntry] = dict()

        # initialize driver entries
        for i in range(1, gen_state.fuzz_driver_number + 1):
            new_entry = cls.DriverEntry(
                id=i,
                is_success=False,
                query_count=0,
                target_functions=[],
            )
            driver_entries[i] = new_entry

        # get query count
        for driver_id, query_count in gen_state.fuzz_driver_query_count.items():
            driver_entries[driver_id].query_count = query_count

        # get is_success
        for driver_id, is_success in gen_state.fuzz_driver_success.items():
            driver_entries[driver_id].is_success = is_success

        # get target functions from fuzz driver file comments
        suffix = (
            ".c"
            if global_vars.library_language == global_vars.SupportedLanguages.C
            else ".cpp"
        )
        for driver_source in driver_out_path.glob(f"fuzz_driver_*{suffix}"):
            # The beginning of driver_source is like:
            # // This fuzz driver is generated for library ngiflib, aiming to fuzz the following functions:
            # // SDL_LoadAnimatedGif at ngiflibSDL.c:103:31 in ngiflibSDL.h
            # // SDL_LoadGIF at ngiflibSDL.c:9:15 in ngiflibSDL.h

            source_name = driver_source.name.removesuffix(suffix)
            driver_id = int(parse("fuzz_driver_{}", source_name)[0])
            if driver_id not in driver_entries:
                logger.warning(
                    f"Fuzz driver {driver_id} in {driver_source} is not found in the generation state."
                )
                continue

            with open(driver_source, "r") as f:
                f.readline()
                while True:
                    line = f.readline()
                    if not line.startswith("//"):
                        break
                    else:
                        func = line.removeprefix("//").strip()
                        driver_entries[driver_id].target_functions.append(func)

        return cls(
            list(driver_entries.values()),
            gen_state.llm_statistics.total_tokens,
            gen_state.duration,
        )


class CrashReport:
    """
    ASan crash report.
    """

    @dataclass
    class CrashEntry:
        """
        An ASan crash data entry.
        """

        signature: str
        """
        The crash signature.
        """

        count: int
        """
        The number of crashes.
        """

        is_learned: bool
        """
        Whether the crash is learned.
        """

        asan_message: str
        """
        The ASan message.

        """

    ILLEGAL_CHARACTERS_RE = re.compile(r"[\000-\010]|[\013-\014]|[\016-\037]")

    @staticmethod
    def filter_excel_string(s: str) -> str:
        """
        Excel has a limit of 32767 characters for a cell, and some illegal characters.
        """
        s = s[-32767:]
        return CrashReport.ILLEGAL_CHARACTERS_RE.sub("", s)

    def add_to_workbook(self, workbook: Workbook):
        """
        Add the crash report to a workbook as a new sheet.

        :param workbook: The workbook.
        """
        # Create a new sheet
        crash_sheet = workbook.create_sheet("Crashes")

        # Write the header
        (
            crash_sheet["A1"],
            crash_sheet["B1"],
            crash_sheet["C1"],
            crash_sheet["D1"],
        ) = (
            "Signature",
            "Count",
            "isLearned",
            "ASan Message",
        )
        # bold the header row
        bold = Font(bold=True)
        for cell in crash_sheet["1:1"]:
            cell.font = bold

        # Write the data
        for i, row in enumerate(self.entries, start=2):
            (
                crash_sheet[f"A{i}"],
                crash_sheet[f"B{i}"],
                crash_sheet[f"C{i}"],
                crash_sheet[f"D{i}"],
            ) = (
                row.signature,
                row.count,
                row.is_learned,
                self.filter_excel_string(row.asan_message),
            )

        # set styles
        # bold the header row
        bold = Font(bold=True, name="Arial")
        for cell in crash_sheet["1:1"]:
            cell.font = bold
        # bold the A column
        for cell in crash_sheet["A:A"]:
            cell.font = bold
        # set column width
        width = {
            "A": 100,
            "B": 7,
            "C": 10,
            "D": 100,
        }
        for col, w in width.items():
            crash_sheet.column_dimensions[col].width = w
        # auto wrap D column
        wrap = Alignment(wrap_text=True)
        for cell in crash_sheet["D:D"]:
            cell.alignment = wrap

    def __init__(self, crash_entries: list[CrashEntry]):
        """
        Initialize the crash report.

        :param crash_entries: The crash entries.
        """
        self.entries = crash_entries

    @classmethod
    def from_collected_data(
        cls,
        gen_state: GenerationState,
    ):
        """
        Create a crash report from collected data.

        :param gen_state: The generation state.
        """
        crash_entries = list()

        for crash_signature, crash_record in gen_state.crashes.items():
            crash_entries.append(
                cls.CrashEntry(
                    signature=crash_signature,
                    count=crash_record.count,
                    is_learned=crash_record.learning_status == "learned",
                    asan_message=crash_record.crashes[0].err_msg,
                )
            )

        return cls(crash_entries)


class StatsReport:
    """
    Class to synthesize and store a statistics report.
    """

    def __init__(
        self,
        state: GenerationState,
        api_collection: APICollection,
        driver_out_path: Path,
    ):
        """
        Initialize the statistics report.
        """
        self.state = state
        self.api_collection = api_collection
        self.driver_out_path = driver_out_path

    @cached_property
    def function_report(self) -> FunctionReport:
        """
        The function report.
        """
        return FunctionReport.from_collected_data(self.api_collection, self.state)

    @cached_property
    def driver_report(self) -> DriverReport:
        """
        The driver report.
        """
        return DriverReport.from_collected_data(self.state, self.driver_out_path)

    @cached_property
    def crash_report(self) -> CrashReport:
        """
        The crash report.
        """
        return CrashReport.from_collected_data(self.state)

    def save_report(self, out_path: Path):
        """
        Save the statistics report to a file.

        :param out_path: The output path.
        """
        report = Workbook()

        # remove the default sheet
        report.remove(report.active)

        self.function_report.add_to_workbook(report)
        self.driver_report.add_to_workbook(report)
        self.crash_report.add_to_workbook(report)

        report.save(out_path)
